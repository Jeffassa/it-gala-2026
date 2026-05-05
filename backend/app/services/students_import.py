"""Parse an Excel file (xlsx) and ingest ESATIC students into the directory.

Smart behavior :
- Reads `nom` and `prenom` separately when both are present, then concatenates
  as full_name = "PRENOM NOM". If only `nom` (or `nom complet`) is provided,
  uses it as-is.
- Reads `genre` / `sexe` and normalizes to "M" or "F".
- If neither a `classe` column nor a `default_classe` is provided, the sheet
  name is used as the classe (typical ESATIC layout: one class per sheet).
- Promotion priority: column → form default → sheet name (if no classe-fallback).
"""
from __future__ import annotations

import io
import re

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.student import Student


MATRICULE_RE = re.compile(r"^\d{2}-ESATIC\d{4}[A-Z]{2}$")


# Tolerant header mapping: Excel header (lowercased / accents stripped) -> internal field
HEADER_MAP = {
    # Matricule
    "matricule": "matricule",
    "matricul": "matricule",
    "matr": "matricule",
    "matricule etudiant": "matricule",

    # Last name only
    "nom": "last_name",
    "noms": "last_name",
    "nom de famille": "last_name",

    # First name only
    "prenom": "first_name",
    "prenoms": "first_name",

    # Already-combined full name (legacy support)
    "nom complet": "full_name",
    "prenom et nom": "full_name",
    "nom et prenom": "full_name",
    "nom & prenom": "full_name",
    "etudiant": "full_name",

    # Gender
    "genre": "gender",
    "sexe": "gender",
    "sex": "gender",
    "m/f": "gender",

    # Email
    "email": "email",
    "e-mail": "email",
    "mail": "email",
    "courriel": "email",

    # Promotion
    "promotion": "promotion",
    "promo": "promotion",
    "annee": "promotion",

    # Classe
    "classe": "classe",
    "filiere": "classe",
    "specialite": "classe",
    "groupe": "classe",

    # Phone
    "telephone": "phone",
    "tel": "phone",
    "phone": "phone",
    "tel.": "phone",
    "numero": "phone",
}


def _normalize_header(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    for a, b in [("é", "e"), ("è", "e"), ("ê", "e"), ("à", "a"), ("â", "a"), ("ô", "o"), ("ç", "c"), ("î", "i"), ("ï", "i"), ("ù", "u"), ("û", "u")]:
        s = s.replace(a, b)
    return s


def _stringify(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _normalize_gender(raw: str) -> str | None:
    """Accepts: M, F, m, f, masculin, feminin, homme, femme, garcon, fille, H, h."""
    if not raw:
        return None
    s = _normalize_header(raw)  # lowercase + strip accents
    if not s:
        return None
    if s[0] in ("m", "h"):  # masculin / homme / m
        return "M"
    if s[0] in ("f",):  # feminin / femme / fille / f
        return "F"
    return None  # silently ignored if unrecognized


def import_students_from_xlsx(
    db: Session,
    content: bytes,
    default_promotion: str | None = None,
    default_classe: str | None = None,
) -> dict:
    """Parse an .xlsx and upsert students. Returns a stats dict.

    The first non-empty row that contains at least one known header is used as header row.
    Subsequent non-empty rows are treated as data.
    """
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    created = 0
    updated = 0
    skipped = 0
    errors: list[str] = []
    total = 0

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find header row: the first row in the first 10 that maps a matricule
        # column AND at least one name column (last_name OR full_name).
        header_idx = -1
        header_map: dict[int, str] = {}
        for ri, row in enumerate(rows[:10]):
            tentative: dict[int, str] = {}
            for ci, val in enumerate(row):
                key = HEADER_MAP.get(_normalize_header(_stringify(val)))
                if key:
                    tentative[ci] = key
            fields = set(tentative.values())
            if "matricule" in fields and ("last_name" in fields or "full_name" in fields):
                header_idx = ri
                header_map = tentative
                break

        if header_idx == -1:
            errors.append(
                f"Feuille « {ws.title} » : entêtes 'matricule' et 'nom' introuvables. Ignorée."
            )
            continue

        # Per-sheet defaults: if no classe column was found in this sheet,
        # the sheet title is used as the classe (ESATIC layout: 1 sheet = 1 classe).
        sheet_has_classe_col = "classe" in set(header_map.values())
        sheet_has_promotion_col = "promotion" in set(header_map.values())

        for ri, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            data: dict[str, str] = {
                f: ""
                for f in ("matricule", "full_name", "first_name", "last_name",
                          "email", "promotion", "classe", "gender", "phone")
            }
            for ci, field in header_map.items():
                if ci < len(row):
                    data[field] = _stringify(row[ci])

            # skip empty rows
            if not any(data.values()):
                continue
            total += 1

            # Matricule
            matricule = data["matricule"].upper().replace(" ", "")
            if not matricule:
                skipped += 1
                errors.append(f"Ligne {ri} (feuille « {ws.title} ») : matricule manquant — ignorée.")
                continue
            if not MATRICULE_RE.match(matricule):
                skipped += 1
                errors.append(f"Ligne {ri} : matricule « {matricule} » invalide (format attendu AA-ESATIC####XX).")
                continue

            # Compute full_name from prenom + nom OR full_name column
            if data["last_name"] or data["first_name"]:
                full_name = f"{data['first_name']} {data['last_name']}".strip()
            else:
                full_name = data["full_name"]
            full_name = " ".join(full_name.split())  # collapse whitespace

            if not full_name:
                skipped += 1
                errors.append(f"Ligne {ri} : nom manquant pour {matricule}.")
                continue

            # Classe : column → form default → sheet name (when no classe column at all)
            if sheet_has_classe_col:
                classe = data["classe"] or default_classe or None
            else:
                classe = data["classe"] or default_classe or ws.title

            # Promotion : column → form default → sheet name (only if sheet name was NOT used as classe)
            if sheet_has_promotion_col:
                promotion = data["promotion"] or default_promotion or ws.title
            elif sheet_has_classe_col:
                # sheet name is free for promotion fallback
                promotion = data["promotion"] or default_promotion or ws.title
            else:
                # sheet name already used for classe -> require explicit promotion
                promotion = data["promotion"] or default_promotion or ""
                if not promotion:
                    skipped += 1
                    errors.append(
                        f"Ligne {ri} : promotion manquante pour {matricule} "
                        f"(le nom de feuille est utilisé pour la classe — précisez une "
                        f"promotion par défaut dans le formulaire d'import)."
                    )
                    continue

            gender = _normalize_gender(data["gender"])
            email = data["email"] or None
            phone = data["phone"] or None

            existing = db.scalar(select(Student).where(Student.matricule == matricule))
            if existing:
                existing.full_name = full_name
                existing.email = email
                existing.promotion = promotion
                existing.classe = classe
                existing.gender = gender
                existing.phone = phone
                updated += 1
            else:
                db.add(Student(
                    matricule=matricule,
                    full_name=full_name,
                    email=email,
                    promotion=promotion,
                    classe=classe,
                    gender=gender,
                    phone=phone,
                ))
                created += 1

    db.commit()
    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:50],
        "total_rows": total,
    }
