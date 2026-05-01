"""Build a styled Excel workbook for the admin report (multi-sheet)."""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.category import Category
from app.models.gala import Gala
from app.models.nominee import Nominee
from app.models.scan import Scan
from app.models.ticket import Ticket, TicketStatus
from app.models.user import User, UserRole
from app.models.vote import Vote


# Brand palette
PRIMARY = "7B0202"
PRIMARY_DEEP = "330808"
ACCENT = "F0A50C"
ACCENT_BRIGHT = "FBC23A"
INK = "0E0808"
INK_LIGHT = "FFFFFF"
LINE = "3A2828"

THIN = Side(style="thin", color=LINE)
BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def _header_style(cell) -> None:
    cell.font = Font(bold=True, color=INK_LIGHT, size=11, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=PRIMARY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BOX


def _kpi_style(cell, accent: bool = False) -> None:
    cell.font = Font(bold=True, size=22, color=ACCENT if accent else INK_LIGHT, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=PRIMARY_DEEP)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BOX


def _label_style(cell) -> None:
    cell.font = Font(size=10, color="C9B8B8", name="Calibri")
    cell.fill = PatternFill("solid", fgColor=PRIMARY_DEEP)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BOX


def _row_style(cell, alt: bool = False) -> None:
    cell.font = Font(size=11, color="000000", name="Calibri")
    cell.fill = PatternFill("solid", fgColor="F8F1E5" if alt else "FFFFFF")
    cell.alignment = Alignment(vertical="center")
    cell.border = BOX


def _title_style(cell) -> None:
    cell.font = Font(bold=True, size=18, color=PRIMARY, name="Calibri")
    cell.alignment = Alignment(horizontal="left", vertical="center")


TICKET_TYPE_LABEL = {"solo": "Solo", "duo": "Duo", "gbonhi": "Gbonhi"}
ROLE_LABEL = {"super_admin": "Super Admin", "cashier": "Caissière", "controller": "Contrôleur", "participant": "Participant"}


def build_report_xlsx(db: Session) -> bytes:
    wb = Workbook()
    active_gala = db.scalar(select(Gala).where(Gala.is_active.is_(True)).order_by(Gala.edition_year.desc()))

    # ============ Sheet 1: Synthèse ============
    ws = wb.active
    ws.title = "Synthèse"

    ws["B2"] = "IT Gala — Rapport d'édition"
    ws["B2"].font = Font(bold=True, size=22, color=PRIMARY, name="Calibri")
    ws["B3"] = active_gala.name + " · " + str(active_gala.edition_year) if active_gala else "—"
    ws["B3"].font = Font(italic=True, size=12, color="555555", name="Calibri")
    ws["B4"] = "Généré le " + datetime.now().strftime("%d/%m/%Y %H:%M")
    ws["B4"].font = Font(size=10, color="888888", name="Calibri")

    total_galas = int(db.scalar(select(func.count(Gala.id))) or 0)
    total_users = int(db.scalar(select(func.count(User.id))) or 0)
    total_participants = int(db.scalar(select(func.count(User.id)).where(User.role == UserRole.PARTICIPANT)) or 0)
    total_tickets = int(db.scalar(select(func.count(Ticket.id))) or 0)
    total_scanned = int(db.scalar(select(func.count(Ticket.id)).where(Ticket.status == TicketStatus.SCANNED)) or 0)
    total_revenue = float(db.scalar(select(func.coalesce(func.sum(Ticket.price), 0))) or 0)
    total_votes = int(db.scalar(select(func.count(Vote.id))) or 0)
    total_categories = int(db.scalar(select(func.count(Category.id))) or 0)
    total_nominees = int(db.scalar(select(func.count(Nominee.id))) or 0)
    total_scans = int(db.scalar(select(func.count(Scan.id))) or 0)

    kpis = [
        ("Recettes totales", f"{int(total_revenue):,}".replace(",", " ") + " FCFA", True),
        ("Tickets vendus", str(total_tickets), False),
        ("Tickets scannés", str(total_scanned), False),
        ("Votes enregistrés", str(total_votes), False),
        ("Participants", str(total_participants), False),
        ("Nominés en lice", str(total_nominees), False),
        ("Catégories", str(total_categories), False),
        ("Total scans", str(total_scans), False),
    ]
    start_row = 6
    for i, (label, value, accent) in enumerate(kpis):
        col_letter = chr(ord("B") + (i % 4) * 2)
        col_next = chr(ord("B") + (i % 4) * 2 + 1)
        row = start_row + (i // 4) * 4
        ws.merge_cells(f"{col_letter}{row}:{col_next}{row}")
        c1 = ws[f"{col_letter}{row}"]
        c1.value = value
        _kpi_style(c1, accent)
        ws.row_dimensions[row].height = 36
        ws.merge_cells(f"{col_letter}{row + 1}:{col_next}{row + 1}")
        c2 = ws[f"{col_letter}{row + 1}"]
        c2.value = label
        _label_style(c2)

    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[col].width = 18

    # ============ Sheet 2: Ventes par type ============
    ws2 = wb.create_sheet("Ventes par type")
    ws2["B2"] = "Ventes par type de ticket"
    _title_style(ws2["B2"])
    headers = ["Type", "Nombre vendu", "Recettes (FCFA)", "Part (%)"]
    for i, h in enumerate(headers):
        c = ws2.cell(row=4, column=i + 2, value=h)
        _header_style(c)
    type_rows = db.execute(
        select(Ticket.type, func.count(Ticket.id), func.coalesce(func.sum(Ticket.price), 0)).group_by(Ticket.type)
    ).all()
    total_count = sum(int(r[1]) for r in type_rows) or 1
    for i, r in enumerate(type_rows):
        type_label = TICKET_TYPE_LABEL.get(str(r[0]), str(r[0]))
        cells = [type_label, int(r[1]), float(r[2]), round(int(r[1]) / total_count * 100, 1)]
        for j, val in enumerate(cells):
            c = ws2.cell(row=5 + i, column=2 + j, value=val)
            _row_style(c, alt=(i % 2 == 1))
            if j == 2:
                c.number_format = "# ##0"
            elif j == 3:
                c.number_format = "0.0\"%\""
    if not type_rows:
        c = ws2.cell(row=5, column=2, value="Aucune vente enregistrée")
        c.font = Font(italic=True, color="888888")
        ws2.merge_cells("B5:E5")

    for col, w in zip(["B", "C", "D", "E"], [22, 18, 22, 14]):
        ws2.column_dimensions[col].width = w

    # ============ Sheet 3: Résultats par catégorie ============
    ws3 = wb.create_sheet("Résultats")
    ws3["B2"] = "Résultats des votes par catégorie"
    _title_style(ws3["B2"])
    headers = ["Catégorie", "Lauréat en tête", "Votes lauréat", "Votes total", "Avance (%)"]
    for i, h in enumerate(headers):
        c = ws3.cell(row=4, column=i + 2, value=h)
        _header_style(c)

    cats = db.scalars(select(Category).order_by(Category.order_index, Category.id)).all()
    for i, cat in enumerate(cats):
        rows = db.execute(
            select(Nominee.id, Nominee.name, func.count(Vote.id).label("votes"))
            .outerjoin(Vote, Vote.nominee_id == Nominee.id)
            .where(Nominee.category_id == cat.id)
            .group_by(Nominee.id, Nominee.name)
            .order_by(func.count(Vote.id).desc())
        ).all()
        total = sum(int(r.votes) for r in rows)
        leader = rows[0] if rows else None
        leader_name = str(leader.name) if leader and int(leader.votes) > 0 else "—"
        leader_votes = int(leader.votes) if leader else 0
        share = round(leader_votes / total * 100, 1) if total else 0.0
        cells = [cat.name, leader_name, leader_votes, total, share]
        for j, val in enumerate(cells):
            c = ws3.cell(row=5 + i, column=2 + j, value=val)
            _row_style(c, alt=(i % 2 == 1))
            if j == 4:
                c.number_format = "0.0\"%\""
    for col, w in zip(["B", "C", "D", "E", "F"], [28, 28, 14, 14, 14]):
        ws3.column_dimensions[col].width = w

    # ============ Sheet 4: Tickets détaillés ============
    ws4 = wb.create_sheet("Tickets")
    ws4["B2"] = "Liste des tickets"
    _title_style(ws4["B2"])
    headers = ["Code", "Type", "Acheteur", "Email", "Téléphone", "Statut", "Prix (FCFA)", "Vendu le", "Scanné le"]
    for i, h in enumerate(headers):
        c = ws4.cell(row=4, column=i + 2, value=h)
        _header_style(c)
    tickets = db.scalars(select(Ticket).order_by(Ticket.sold_at.desc())).all()
    status_label = {"sold": "Vendu", "scanned": "Scanné", "cancelled": "Annulé"}
    for i, t in enumerate(tickets):
        cells = [
            t.code,
            TICKET_TYPE_LABEL.get(str(t.type), str(t.type)),
            t.buyer_full_name,
            t.buyer_email,
            t.buyer_phone or "—",
            status_label.get(str(t.status), str(t.status)),
            float(t.price),
            t.sold_at.strftime("%d/%m/%Y %H:%M") if t.sold_at else "",
            t.scanned_at.strftime("%d/%m/%Y %H:%M") if t.scanned_at else "—",
        ]
        for j, val in enumerate(cells):
            c = ws4.cell(row=5 + i, column=2 + j, value=val)
            _row_style(c, alt=(i % 2 == 1))
            if j == 6:
                c.number_format = "# ##0"
    if not tickets:
        c = ws4.cell(row=5, column=2, value="Aucun ticket")
        c.font = Font(italic=True, color="888888")
    for col, w in zip(["B", "C", "D", "E", "F", "G", "H", "I", "J"], [20, 12, 26, 28, 16, 12, 14, 18, 18]):
        ws4.column_dimensions[col].width = w

    # ============ Sheet 5: Utilisateurs ============
    ws5 = wb.create_sheet("Utilisateurs")
    ws5["B2"] = "Utilisateurs et dépenses"
    _title_style(ws5["B2"])
    headers = ["Nom", "Email", "Rôle", "Statut", "Tickets achetés", "Total dépensé (FCFA)"]
    for i, h in enumerate(headers):
        c = ws5.cell(row=4, column=i + 2, value=h)
        _header_style(c)
    users = db.scalars(select(User).order_by(User.created_at.desc())).all()
    spend_rows = db.execute(
        select(Ticket.buyer_email, func.count(Ticket.id).label("cnt"), func.coalesce(func.sum(Ticket.price), 0).label("tot"))
        .group_by(Ticket.buyer_email)
    ).all()
    spend_map = {row.buyer_email: (int(row.cnt), float(row.tot)) for row in spend_rows}
    for i, u in enumerate(users):
        cnt, tot = spend_map.get(u.email, (0, 0.0))
        cells = [
            u.full_name,
            u.email,
            ROLE_LABEL.get(str(u.role), str(u.role)),
            "Actif" if u.is_active else "Suspendu",
            cnt,
            tot,
        ]
        for j, val in enumerate(cells):
            c = ws5.cell(row=5 + i, column=2 + j, value=val)
            _row_style(c, alt=(i % 2 == 1))
            if j == 5:
                c.number_format = "# ##0"
    for col, w in zip(["B", "C", "D", "E", "F", "G"], [24, 28, 14, 14, 16, 22]):
        ws5.column_dimensions[col].width = w

    # Set first row tall + freeze on each table sheet
    for sheet in [ws2, ws3, ws4, ws5]:
        sheet.row_dimensions[4].height = 26
        sheet.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
